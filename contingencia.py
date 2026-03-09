import streamlit as st
import pandas as pd
import psycopg2
import numpy as np
from datetime import datetime
from io import BytesIO
import re
import concurrent.futures
from pyDolarVenezuela.pages import BCV
from pyDolarVenezuela import Monitor
import pytz

# ============================================================
# ESTILOS CSS — TEMA CLARO PROFESIONAL
# ============================================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&family=Fraunces:ital,wght@0,600;0,700;1,600&display=swap');

html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }

.stApp { background: #f0f2f8; color: #1e293b; }

[data-testid="stSidebar"] {
    background: #ffffff !important;
    border-right: 1px solid #e2e8f0 !important;
    box-shadow: 2px 0 12px rgba(0,0,0,0.05);
}
[data-testid="stSidebar"] .stMarkdown h2 {
    color: #1e293b !important; font-size: 0.72rem !important;
    font-weight: 700 !important; letter-spacing: 0.12em !important;
    text-transform: uppercase !important; margin-bottom: 0.4rem !important;
}
[data-testid="stSidebar"] .stMarkdown h3 {
    color: #475569 !important; font-size: 0.72rem !important;
    font-weight: 600 !important; letter-spacing: 0.10em !important;
    text-transform: uppercase !important; margin-bottom: 0.4rem !important;
}

h1 { font-family: 'Fraunces', serif !important; font-size: 1.9rem !important; font-weight: 700 !important; color: #0f172a !important; letter-spacing: -0.02em !important; }
h2 { color: #1e293b !important; font-weight: 600 !important; }
h3 { color: #334155 !important; font-weight: 600 !important; }
h4, h5 { color: #475569 !important; font-weight: 600 !important; }

[data-testid="stMetric"] {
    background: #ffffff; border: 1px solid #e2e8f0; border-radius: 12px;
    padding: 1rem 1.2rem !important; box-shadow: 0 1px 4px rgba(0,0,0,0.06);
    transition: box-shadow 0.2s, border-color 0.2s;
}
[data-testid="stMetric"]:hover { box-shadow: 0 4px 16px rgba(0,0,0,0.10); border-color: #3b82f6; }
[data-testid="stMetricLabel"] { color: #64748b !important; font-size: 0.75rem !important; font-weight: 600 !important; text-transform: uppercase; letter-spacing: 0.07em; }
[data-testid="stMetricValue"] { color: #0f172a !important; font-size: 1.8rem !important; font-weight: 700 !important; font-family: 'DM Mono', monospace !important; }

.stAlert { border-radius: 10px !important; }
.stSuccess { background: #f0fdf4 !important; border: 1px solid #bbf7d0 !important; border-left: 4px solid #22c55e !important; color: #14532d !important; }
.stWarning { background: #fffbeb !important; border: 1px solid #fde68a !important; border-left: 4px solid #f59e0b !important; color: #78350f !important; }
.stError   { background: #fef2f2 !important; border: 1px solid #fecaca !important; border-left: 4px solid #ef4444 !important; color: #7f1d1d !important; }
.stInfo    { background: #eff6ff !important; border: 1px solid #bfdbfe !important; border-left: 4px solid #3b82f6 !important; color: #1e3a5f !important; }

[data-testid="stTabs"] [role="tablist"] {
    background: #f1f5f9; border-radius: 10px; padding: 4px; gap: 2px; border: 1px solid #e2e8f0;
}
[data-testid="stTabs"] [role="tab"] {
    border-radius: 8px !important; color: #64748b !important; font-size: 0.82rem !important;
    font-weight: 500 !important; padding: 6px 14px !important; transition: all 0.15s !important;
}
[data-testid="stTabs"] [role="tab"][aria-selected="true"] {
    background: #2563eb !important; color: #fff !important;
    box-shadow: 0 2px 8px rgba(37,99,235,0.25) !important;
}

[data-testid="stExpander"] {
    background: #ffffff !important; border: 1px solid #e2e8f0 !important;
    border-radius: 10px !important; box-shadow: 0 1px 3px rgba(0,0,0,0.05) !important;
    margin-bottom: 6px !important;
}
[data-testid="stExpander"] summary { font-weight: 500 !important; color: #334155 !important; font-size: 0.88rem !important; }

[data-testid="stDataFrame"] {
    border-radius: 10px !important; overflow: hidden !important;
    border: 1px solid #e2e8f0 !important; box-shadow: 0 1px 4px rgba(0,0,0,0.05) !important;
}

/* ── FILE UPLOADER REDISEÑADO ── */
[data-testid="stFileUploader"] {
    background: transparent !important;
    border: none !important;
}
[data-testid="stFileUploaderDropzone"] {
    background: #ffffff !important;
    border: 2px dashed #c7d2e8 !important;
    border-radius: 20px !important;
    padding: 2.5rem 2rem !important;
    transition: all 0.25s ease !important;
    box-shadow: 0 4px 24px rgba(37,99,235,0.06), 0 1px 4px rgba(0,0,0,0.04) !important;
    position: relative !important;
}
[data-testid="stFileUploaderDropzone"]:hover {
    border-color: #3b82f6 !important;
    background: #f8faff !important;
    box-shadow: 0 8px 40px rgba(37,99,235,0.12), 0 2px 8px rgba(0,0,0,0.06) !important;
    transform: translateY(-2px) !important;
}
[data-testid="stFileUploaderDropzoneInstructions"] {
    font-size: 0.9rem !important;
    color: #64748b !important;
}

.stDownloadButton > button {
    background: linear-gradient(135deg, #1d4ed8, #4f46e5) !important; color: white !important;
    border: none !important; border-radius: 10px !important; font-weight: 600 !important;
    padding: 0.6rem 1.6rem !important; font-size: 0.9rem !important;
    box-shadow: 0 2px 10px rgba(37,99,235,0.3) !important; width: 100% !important;
}
.stDownloadButton > button:hover { opacity: 0.9 !important; }

hr { border-color: #e2e8f0 !important; margin: 1.5rem 0 !important; }

.expediente-header {
    background: #ffffff; border: 1px solid #e2e8f0; border-radius: 14px;
    padding: 1.4rem 2rem; margin-bottom: 1.5rem; display: flex;
    gap: 2.5rem; flex-wrap: wrap; align-items: center;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
}
.exp-label { font-size: 0.68rem; color: #94a3b8; font-weight: 700; text-transform: uppercase; letter-spacing: 0.1em; }
.exp-value { font-size: 0.95rem; color: #1e293b; font-weight: 600; margin-top: 3px; }

.bcv-badge {
    display: inline-flex; align-items: center; gap: 8px;
    background: #f0fdf4; border: 1px solid #bbf7d0; border-radius: 999px;
    padding: 6px 16px; font-size: 0.83rem; color: #15803d; font-weight: 600; margin-bottom: 1.2rem;
}
.bcv-badge-cache {
    display: inline-flex; align-items: center; gap: 8px;
    background: #fffbeb; border: 1px solid #fde68a; border-radius: 999px;
    padding: 6px 16px; font-size: 0.83rem; color: #92400e; font-weight: 600; margin-bottom: 1.2rem;
}

/* ── UPLOAD HERO CARD ── */
.upload-hero {
    background: #ffffff;
    border-radius: 24px;
    border: 1px solid #e2e8f0;
    padding: 0;
    overflow: hidden;
    box-shadow: 0 4px 32px rgba(37,99,235,0.08), 0 1px 4px rgba(0,0,0,0.04);
    margin-top: 0.5rem;
}
.upload-hero-top {
    background: linear-gradient(135deg, #1e3a8a 0%, #1d4ed8 50%, #4f46e5 100%);
    padding: 2.2rem 2.5rem 2rem;
    position: relative;
    overflow: hidden;
}
.upload-hero-top::before {
    content: '';
    position: absolute;
    top: -40px; right: -40px;
    width: 200px; height: 200px;
    background: rgba(255,255,255,0.06);
    border-radius: 50%;
}
.upload-hero-top::after {
    content: '';
    position: absolute;
    bottom: -60px; left: 60px;
    width: 160px; height: 160px;
    background: rgba(255,255,255,0.04);
    border-radius: 50%;
}
.upload-hero-title {
    font-family: 'Fraunces', serif;
    font-size: 1.5rem;
    font-weight: 700;
    color: #ffffff;
    margin: 0 0 0.4rem;
    letter-spacing: -0.02em;
    position: relative; z-index: 1;
}
.upload-hero-sub {
    font-size: 0.85rem;
    color: rgba(255,255,255,0.7);
    margin: 0;
    position: relative; z-index: 1;
}
.upload-hero-body {
    padding: 1.8rem 2.5rem 2rem;
}
.upload-chips {
    display: flex;
    gap: 0.6rem;
    flex-wrap: wrap;
    margin-bottom: 1.4rem;
}
.upload-chip {
    display: inline-flex; align-items: center; gap: 6px;
    background: #f1f5f9; border: 1px solid #e2e8f0;
    border-radius: 999px; padding: 5px 13px;
    font-size: 0.76rem; color: #475569; font-weight: 600;
}
.upload-chip span { font-size: 0.95rem; }

::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #f1f5f9; }
::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #94a3b8; }
</style>
""", unsafe_allow_html=True)


# ============================================================
# BASE DE DATOS
# ============================================================
def conectar_contingencia():
    try:
        conexion = psycopg2.connect(
            host=st.secrets["supabase"]["host"],
            port=st.secrets["supabase"]["port"],
            database=st.secrets["supabase"]["database"],
            user=st.secrets["supabase"]["user"],
            password=st.secrets["supabase"]["password"]
        )
        return conexion
    except Exception as e:
        st.error(f"Fallo la conexión: {e}")
        return None

conn = conectar_contingencia()
if conn:
    st.success("✅ Conectado exitosamente")
    conn.close()


# ============================================================
# TASA BCV — USD + EUR, TIMEOUT 30s, FALLBACK A BD
# ============================================================

def _crear_tabla_tasa_si_no_existe(conn):
    """Crea la tabla de historial de tasas BCV si no existe aún."""
    try:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS tasa_bcv_historico (
                    id          SERIAL PRIMARY KEY,
                    moneda      VARCHAR(10)    NOT NULL DEFAULT 'EUR',
                    tasa        NUMERIC(18, 4) NOT NULL,
                    fecha_bcv   VARCHAR(60),
                    registrado  TIMESTAMPTZ    NOT NULL DEFAULT NOW()
                );
            """)
            conn.commit()
    except Exception:
        conn.rollback()


def _guardar_tasa_en_db(tasa: float, fecha_str: str, moneda: str):
    """Persiste la tasa recién obtenida del BCV en la tabla histórica."""
    try:
        conn = conectar_contingencia()
        if conn is None:
            return
        _crear_tabla_tasa_si_no_existe(conn)
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO tasa_bcv_historico (moneda, tasa, fecha_bcv) VALUES (%s, %s, %s)",
                (moneda, tasa, fecha_str)
            )
            conn.commit()
        conn.close()
    except Exception:
        pass  # No bloquear la app si falla el guardado


def _obtener_ultima_tasa_db(moneda: str):
    """
    Recupera el registro más reciente de la tabla para la moneda indicada.
    Retorna (tasa_float, label_str) o (None, None) si no hay datos.
    """
    try:
        conn = conectar_contingencia()
        if conn is None:
            return None, None
        _crear_tabla_tasa_si_no_existe(conn)
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT tasa, fecha_bcv, registrado
                FROM tasa_bcv_historico
                WHERE moneda = %s
                ORDER BY registrado DESC, id DESC
                LIMIT 1
                """,
                (moneda,)
            )
            row = cur.fetchone()
        conn.close()
        if row:
            tasa_db, fecha_bcv_db, registrado_db = row
            zone = pytz.timezone('America/Caracas')
            try:
                registrado_str = registrado_db.astimezone(zone).strftime('%d/%m/%Y, %I:%M %p')
            except Exception:
                registrado_str = str(registrado_db)
            label = f"{fecha_bcv_db or registrado_str} ⚠️ (caché BD)"
            return float(tasa_db), label
        return None, None
    except Exception:
        return None, None


@st.cache_data(ttl=3600)
def obtener_tasas_bcv():
    """
    Obtiene las tasas USD y EUR del BCV con timeout de 30 s por moneda.
    Si falla, usa el último dato guardado en BD sin mostrar mensajes molestos.
    Si no hay nada en BD, retorna tasa=None silenciosamente.
    """
    zone = pytz.timezone('America/Caracas')

    def _fetch_moneda(moneda_key: str):
        monitor = Monitor(BCV, 'USD')
        data = monitor.get_value_monitors(moneda_key.lower())
        tasa = float(data.price)
        fecha_str = data.last_update.astimezone(zone).strftime('%d/%m/%Y, %I:%M %p')
        return tasa, fecha_str

    resultado = {}
    errores = []

    for moneda in ("USD", "EUR"):
        tasa, fecha_str, es_cache = None, "Sin datos disponibles", False

        try:
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
                future = executor.submit(_fetch_moneda, moneda)
                tasa, fecha_str = future.result(timeout=30)
        except Exception:
            es_cache = True

        if tasa is not None:
            # BCV respondió: guardar dato fresco
            _guardar_tasa_en_db(tasa, fecha_str, moneda)
        else:
            # Intentar fallback en BD
            tasa_db, label_db = _obtener_ultima_tasa_db(moneda)
            if tasa_db is not None:
                tasa, fecha_str = tasa_db, label_db
                es_cache = True
            else:
                errores.append(moneda)

        resultado[moneda] = {"tasa": tasa, "fecha": fecha_str, "es_cache": es_cache}

    # Solo advertir si EUR (la tasa que se usa en cálculos) no está disponible
    eur = resultado.get("EUR", {})
    if eur.get("tasa") is None:
        st.warning(
            "⚠️ No se pudo obtener la tasa **EUR** del BCV y tampoco hay datos guardados en la BD. "
            "Ejecuta el SQL de inicialización en Supabase para activar el fallback."
        )
    elif eur.get("es_cache"):
        st.info("🗄️ Tasa BCV EUR cargada desde caché (BD). El sitio del BCV no respondió.")

    return resultado


@st.cache_data(ttl=600)
def obtener_calendario_db():
    try:
        conn = conectar_contingencia()
        if conn is None:
            return pd.DataFrame()
        query = """
            SELECT c.digito_rif, r.codigo_renta, c.anio, c.mes,
                   COALESCE(c.quincena, 'Mensual') AS quincena,
                   c.fecha_limite, tc.tipo AS tipo_contribuyente,
                   COALESCE(c.tipo_cierre, 'Ordinario') AS tipo_cierre
            FROM calendario_seniat c
            JOIN dim_renta r ON c.id_renta = r.id_renta
            JOIN dim_tipo_contribuyente tc ON c.id_tipo_contribuyente = tc.id_tipo_contribuyente;
        """
        df_cal = pd.read_sql_query(query, conn)
        conn.close()
        df_cal['fecha_limite'] = pd.to_datetime(df_cal['fecha_limite'])
        df_cal['anio'] = df_cal['anio'].astype(int)
        df_cal['mes'] = df_cal['mes'].astype(int)
        df_cal['digito_rif'] = df_cal['digito_rif'].astype(int)
        df_cal['quincena'] = df_cal['quincena'].fillna('Mensual').replace({'': 'Mensual', 'None': 'Mensual'})
        df_cal['tipo_contribuyente'] = df_cal['tipo_contribuyente'].astype(str).str.strip().str.lower()
        df_cal['tipo_cierre'] = df_cal['tipo_cierre'].astype(str).str.strip().str.lower()
        return df_cal
    except Exception as e:
        st.error(f"Error al obtener el calendario: {e}")
        return pd.DataFrame()


# ============================================================
# LÓGICA PRINCIPAL
# ============================================================
def parsear_fecha_operacion(val):
    try:
        if pd.isna(val):
            return pd.NaT
    except Exception:
        pass

    if isinstance(val, np.datetime64):
        try:
            return pd.Timestamp(val)
        except Exception:
            return pd.NaT

    if isinstance(val, (datetime, pd.Timestamp)):
        return pd.Timestamp(val)

    if isinstance(val, (int, float)):
        try:
            return pd.Timestamp('1899-12-30') + pd.Timedelta(days=int(val))
        except Exception:
            return pd.NaT

    val_str = str(val).strip()
    if not val_str or val_str.lower() in ('nan', 'none', 'nat', 'nd', '-', ''):
        return pd.NaT

    val_str = val_str.split(' ')[0].strip()

    for fmt in (
        '%d/%m/%Y',
        '%d-%m-%Y',
        '%Y-%m-%d',
        '%d/%m/%y',
        '%m/%d/%Y',
        '%Y/%m/%d',
    ):
        try:
            return pd.to_datetime(val_str, format=fmt)
        except Exception:
            pass

    return pd.to_datetime(val_str, errors='coerce', dayfirst=True)


def parse_periodo(val):
    if pd.isna(val):
        return 0, 0
    if isinstance(val, np.datetime64):
        try:
            ts = pd.Timestamp(val)
            return ts.month, ts.year
        except Exception:
            return 0, 0
    if isinstance(val, (datetime, pd.Timestamp)):
        return val.month, val.year
    val_str = str(val).lower().strip()

    match_ddmmyyyy = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](20\d{2})', val_str)
    if match_ddmmyyyy:
        return int(match_ddmmyyyy.group(2)), int(match_ddmmyyyy.group(3))

    match_mmyyyy = re.search(r'(\d{1,2})[/\-](20\d{2})', val_str)
    if match_mmyyyy:
        return int(match_mmyyyy.group(1)), int(match_mmyyyy.group(2))

    meses_map = {'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
                 'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12}
    mes, anio = 0, 0
    match_anio = re.search(r'(20\d{2})', val_str)
    if match_anio:
        anio = int(match_anio.group(1))
    else:
        match_anio_2 = re.search(r'(?<!\d)(\d{2})(?!\d)', val_str)
        if match_anio_2:
            anio = 2000 + int(match_anio_2.group(1))
    for m_str, m_num in meses_map.items():
        if m_str in val_str:
            mes = m_num
            break
    if mes == 0:
        match_mes = re.search(r'^(0?[1-9]|1[0-2])[-/]', val_str)
        if match_mes:
            mes = int(match_mes.group(1))
    return mes, anio


def procesar_auditoria_completa(df_ops, df_cal, digito_usuario, fecha_contingencia, tipo_cierre_usuario):
    fecha_contingencia_dt = pd.to_datetime(fecha_contingencia)
    anio_fin_calendario = fecha_contingencia_dt.year

    df_ops.columns = df_ops.columns.str.strip()

    col_doc = [c for c in df_ops.columns if 'documento' in c.lower() or 'planilla' in c.lower() or 'certificado' in c.lower()]
    if not col_doc:
        col_doc = [c for c in df_ops.columns if 'nro' in c.lower() or 'num' in c.lower()]

    duplicados_eliminados = 0

    col_fecha = [c for c in df_ops.columns if 'fecha' in c.lower() and 'operaci' in c.lower()]
    if not col_fecha:
        st.error("❌ No encontré la columna de fecha de operación en el Excel.")
        st.stop()
    nombre_col_fecha = col_fecha[0]

    nombre_concepto = 'Concepto Contable' if 'Concepto Contable' in df_ops.columns else 'Descripción'
    if nombre_concepto not in df_ops.columns:
        df_ops[nombre_concepto] = '---'

    df_ops['Periodo_Original'] = df_ops['Periodo'].astype(str)
    df_ops['Es_Sustitutiva'] = df_ops.astype(str).apply(lambda col: col.str.lower().str.contains('sustitutiva')).any(axis=1)

    df_ops['Fecha_Operacion_Real'] = df_ops[nombre_col_fecha].apply(parsear_fecha_operacion)

    periodos_parsed = df_ops['Periodo'].apply(parse_periodo)
    df_ops['Mes_Num'] = periodos_parsed.apply(lambda x: x[0])
    df_ops['Anio_Num'] = periodos_parsed.apply(lambda x: x[1])
    df_ops['Digito'] = digito_usuario

    df_validos = df_ops[df_ops['Anio_Num'] > 2000].copy()
    if not df_validos.empty:
        df_validos['YYYYMM'] = df_validos['Anio_Num'] * 100 + df_validos['Mes_Num']
        min_ym = df_validos['YYYYMM'].min()
        max_ym = df_validos['YYYYMM'].max()
        mes_ini, anio_ini = int(str(min_ym)[-2:]), int(str(min_ym)[:4])
        mes_fin, anio_fin_excel = int(str(max_ym)[-2:]), int(str(max_ym)[:4])
        periodo_str = f"{mes_ini:02d}/{anio_ini} → {mes_fin:02d}/{anio_fin_excel}"
        anio_inicio = anio_ini
    else:
        periodo_str = "Sin periodos válidos detectados"
        anio_inicio = anio_fin_calendario - 3
        min_ym = anio_inicio * 100 + 1

    def normalizar_ops(row):
        renta = str(row.get('Renta', '')).upper().strip()
        concepto = str(row.get(nombre_concepto, '')).upper().strip()
        if 'ANTICIPO' in renta or 'ANTICIPO' in concepto: return 'ANTICIPO-ISLR'
        if 'IVA' in renta:
            if '30' in renta: return 'IVA/30'
            if '35' in renta: return 'IVA/35'
            return 'IVA/30'
        if 'RET' in renta or 'RET' in concepto: return 'ISLR-RETENCION'
        elif 'ISLR' in renta or 'ISLR' in concepto: return 'ISLR-ANUAL'
        return renta

    df_ops['Renta_Norm'] = df_ops.apply(normalizar_ops, axis=1)

    def es_registro_valido(row):
        concepto_limpio = str(row.get(nombre_concepto, '')).lower().strip()
        descripcion_limpia = str(row.get('Descripción', row.get('descripcion', ''))).lower().strip()
        texto_completo = concepto_limpio + ' ' + descripcion_limpia
        if row['Renta_Norm'] == 'IVA/30':
            if 'credito' in texto_completo or 'crédito' in texto_completo:
                return False
            return True
        if row['Renta_Norm'] in ['ISLR-ANUAL', 'ISLR-RETENCION']:
            if 'sustitutiva' in texto_completo:
                return False
            return True
        return True

    df_ops = df_ops[df_ops.apply(es_registro_valido, axis=1)].copy()
    df_ops['Anio_Cruce'] = df_ops['Anio_Num']
    df_ops['Mes_Cruce'] = df_ops.apply(lambda r: 99 if r['Renta_Norm'] == 'ISLR-ANUAL' else r['Mes_Num'], axis=1)
    df_ops = df_ops[
        (df_ops['Fecha_Operacion_Real'].isna()) | (df_ops['Fecha_Operacion_Real'] <= fecha_contingencia_dt)
    ].copy()

    anticipos_por_mes = {}
    if not df_ops.empty:
        df_ops['YYYYMM'] = df_ops['Anio_Cruce'] * 100 + df_ops['Mes_Cruce']

        candidatos_transicion = []
        df_esp_iva35 = df_ops[df_ops['Renta_Norm'] == 'IVA/35']
        df_esp_anticipo = df_ops[df_ops['Renta_Norm'] == 'ANTICIPO-ISLR']
        if not df_esp_iva35.empty:
            candidatos_transicion.append(df_esp_iva35['YYYYMM'].min())
        if not df_esp_anticipo.empty:
            candidatos_transicion.append(df_esp_anticipo['YYYYMM'].min())

        if candidatos_transicion:
            trans_yyyymm = min(candidatos_transicion)
            df_ops['Tipo_Contribuyente'] = np.where(df_ops['YYYYMM'] >= trans_yyyymm, 'especial', 'ordinario')
            min_general_yyyymm = df_ops['YYYYMM'].min()
            if trans_yyyymm > min_general_yyyymm:
                mes_t = str(trans_yyyymm)[-2:]
                anio_t = str(trans_yyyymm)[:4]
                tipo_final_empresa = f"Transición a Especial ({mes_t}/{anio_t})"
            else:
                tipo_final_empresa = "Especial"
        else:
            df_ops['Tipo_Contribuyente'] = 'ordinario'
            trans_yyyymm = 999999
            tipo_final_empresa = "Ordinario"

        df_ops = df_ops.sort_values(
            by=['Digito', 'Renta_Norm', 'Anio_Cruce', 'Mes_Cruce', 'Fecha_Operacion_Real'],
            ascending=True,
            na_position='last'
        )
        df_ops['Orden_Declaracion'] = df_ops.groupby(
            ['Digito', 'Renta_Norm', 'Anio_Cruce', 'Mes_Cruce']
        ).cumcount() + 1

        anticipos_ops = df_ops[df_ops['Renta_Norm'] == 'ANTICIPO-ISLR']
        anticipos_por_mes = anticipos_ops.groupby(['Anio_Cruce', 'Mes_Cruce']).size().to_dict()

        col_quin = [c for c in df_ops.columns if 'quincena' in c.lower()]
        df_ops['Quincena_Original'] = df_ops[col_quin[0]].apply(
            lambda val: 'Primera' if '1' in str(val) or 'PRIMER' in str(val).upper()
            else ('Segunda' if '2' in str(val) or 'SEGUND' in str(val).upper() else '')
        ) if col_quin else ''

        def deducir_quincena(row):
            if row['Renta_Norm'] == 'ISLR-ANUAL':
                return 'Anual'
            if row['Renta_Norm'] == 'ISLR-RETENCION':
                return 'Mensual'
            if row['Tipo_Contribuyente'] != 'especial':
                return 'Mensual'
            if row['Renta_Norm'] == 'ANTICIPO-ISLR':
                if pd.notna(row['Fecha_Operacion_Real']):
                    return 'Segunda' if 1 <= row['Fecha_Operacion_Real'].day <= 15 else 'Primera'
                if row['Quincena_Original'] in ['Primera', 'Segunda']:
                    return row['Quincena_Original']
                return 'Primera' if row['Orden_Declaracion'] == 1 else 'Segunda'
            elif row['Renta_Norm'] in ['IVA/30', 'IVA/35']:
                if row['Quincena_Original'] in ['Primera', 'Segunda']:
                    return row['Quincena_Original']
                if pd.notna(row.get('fecha_limite')):
                    return 'Primera' if row['fecha_limite'].day <= 20 else 'Segunda'
                return 'Primera' if row['Orden_Declaracion'] == 1 else 'Segunda'
            return 'Mensual'

        df_ops['Quincena_Deducida'] = df_ops.apply(deducir_quincena, axis=1)
    else:
        trans_yyyymm = 999999
        tipo_final_empresa = "Evaluación Sin Pagos Registrados"

    anio_fin = anio_fin_calendario
    anios_presentes = list(range(anio_inicio, anio_fin + 1))
    tipo_c_user = tipo_cierre_usuario.strip().lower()

    def normalizar_cal(row):
        renta = str(row.get('codigo_renta', '')).upper().strip()
        if 'ANTICIPO' in renta: return 'ANTICIPO-ISLR'
        if 'IVA' in renta:
            if '30' in renta: return 'IVA/30'
            if '35' in renta: return 'IVA/35'
            return 'IVA/30'
        if 'RET' in renta: return 'ISLR-RETENCION'
        if '961' in renta or 'ISLR' in renta:
            diff_meses = (row['fecha_limite'].month - row['mes']) % 12
            return 'ISLR-ANUAL' if diff_meses >= 2 else 'ISLR-RETENCION'
        return renta

    df_cal['Renta_Norm'] = df_cal.apply(normalizar_cal, axis=1)

    def get_cal_anio_cruce(row):
        if row['Renta_Norm'] == 'ISLR-ANUAL':
            if row['mes'] >= 10 and row['fecha_limite'].month <= 6: return row['fecha_limite'].year - 1
            return row['fecha_limite'].year
        if row['mes'] >= 10 and row['fecha_limite'].month <= 5: return row['fecha_limite'].year - 1
        return row['fecha_limite'].year

    df_cal['Anio_Cruce'] = df_cal.apply(get_cal_anio_cruce, axis=1)
    df_cal['Mes_Cruce'] = df_cal.apply(lambda r: 99 if r['Renta_Norm'] == 'ISLR-ANUAL' else r['mes'], axis=1)
    df_cal['YYYYMM_cal'] = df_cal['Anio_Cruce'] * 100 + df_cal['mes']
    df_cal['Tipo_Contribuyente_Esperado'] = np.where(df_cal['YYYYMM_cal'] >= trans_yyyymm, 'especial', 'ordinario')

    df_cal_empresa = df_cal[df_cal['digito_rif'] == digito_usuario].copy()
    mask_islr_anual = df_cal_empresa['Renta_Norm'] == 'ISLR-ANUAL'
    mask_cierre_exacto = df_cal_empresa['tipo_cierre'] == tipo_c_user

    if tipo_c_user == 'ordinario':
        df_cal_empresa = df_cal_empresa[~mask_islr_anual | (mask_cierre_exacto & (df_cal_empresa['mes'] == 12))]
    else:
        df_cal_empresa = df_cal_empresa[~mask_islr_anual | mask_cierre_exacto]

    def validar_universo(row):
        t = row['Tipo_Contribuyente_Esperado']
        r = row['Renta_Norm']
        if t == 'especial': return r in ['IVA/30', 'IVA/35', 'ANTICIPO-ISLR', 'ISLR-ANUAL', 'ISLR-RETENCION']
        else: return r in ['IVA/30', 'ISLR-ANUAL', 'ISLR-RETENCION']

    def es_obligacion_valida_desde_inicio(row):
        if row['Renta_Norm'] == 'ISLR-ANUAL': return row['Anio_Cruce'] >= anio_inicio
        else: return row['YYYYMM_cal'] >= min_ym

    df_cal_empresa = df_cal_empresa[
        (df_cal_empresa['Anio_Cruce'] <= anio_fin) &
        (df_cal_empresa.apply(es_obligacion_valida_desde_inicio, axis=1)) &
        (df_cal_empresa['tipo_contribuyente'] == df_cal_empresa['Tipo_Contribuyente_Esperado']) &
        (df_cal_empresa.apply(validar_universo, axis=1))
    ].copy()

    df_cal_empresa = df_cal_empresa[df_cal_empresa['fecha_limite'] <= fecha_contingencia_dt].copy()
    df_cal_empresa.loc[df_cal_empresa['Renta_Norm'] == 'ISLR-ANUAL', 'quincena'] = 'Anual'
    df_cal_empresa.loc[df_cal_empresa['Renta_Norm'] == 'ISLR-RETENCION', 'quincena'] = 'Mensual'

    mask_ord_iva = (df_cal_empresa['Tipo_Contribuyente_Esperado'] == 'ordinario') & (df_cal_empresa['Renta_Norm'] == 'IVA/30')
    df_cal_empresa.loc[mask_ord_iva, 'quincena'] = 'Mensual'

    mask_fantasmas_esp = (
        (df_cal_empresa['Tipo_Contribuyente_Esperado'] == 'especial') &
        (df_cal_empresa['Renta_Norm'].isin(['IVA/30', 'IVA/35', 'ANTICIPO-ISLR'])) &
        (df_cal_empresa['quincena'] == 'Mensual')
    )
    df_cal_empresa = df_cal_empresa[~mask_fantasmas_esp].copy()
    df_cal_empresa = df_cal_empresa.sort_values('fecha_limite').drop_duplicates(
        subset=['Anio_Cruce', 'Mes_Cruce', 'Renta_Norm', 'quincena', 'Tipo_Contribuyente_Esperado'], keep='last'
    )

    if df_ops.empty:
        df_ops_cruce = pd.DataFrame(columns=['Digito', 'Renta_Norm', 'Anio_Cruce', 'Mes_Cruce', 'Quincena_Deducida'])
    else:
        df_ops_cruce = df_ops

    df_audit = pd.merge(
        df_cal_empresa, df_ops_cruce,
        left_on=['digito_rif', 'Renta_Norm', 'Anio_Cruce', 'Mes_Cruce', 'quincena'],
        right_on=['Digito', 'Renta_Norm', 'Anio_Cruce', 'Mes_Cruce', 'Quincena_Deducida'],
        how='outer', indicator=True
    )

    df_audit['Fase_Fiscal'] = df_audit['Tipo_Contribuyente_Esperado'].fillna(df_audit['Tipo_Contribuyente']).str.capitalize()
    df_audit['Frecuencia'] = df_audit['quincena'].fillna(df_audit['Quincena_Deducida'])

    def determinar_estatus(row):
        if pd.notna(row.get('Es_Sustitutiva')) and row.get('Es_Sustitutiva'):
            return '🟢 Exonerado (Sustitutiva)'
        if row['_merge'] == 'left_only':
            if row['Renta_Norm'] == 'ANTICIPO-ISLR':
                if anticipos_por_mes.get((row['Anio_Cruce'], row['Mes_Cruce']), 0) >= 1:
                    return '🟢 Exonerado (Cubierto por Pago Único)'
            dias_hoy = (fecha_contingencia_dt - row['fecha_limite']).days
            return '❌ NO PAGADO (Vencido)' if dias_hoy > 0 else '⏳ Pendiente (Por Vencer)'
        elif row['_merge'] == 'right_only':
            return '⚠️ Pago Extra / Sin Regla'
        else:
            dias = (row['Fecha_Operacion_Real'] - row['fecha_limite']).days
            if dias > 0: estatus_base = '🔴 Pagado Tarde'
            elif dias < 0: estatus_base = '🟢 Pagado Adelantado'
            else: estatus_base = '🔵 Pagado Exacto'
            if row['Renta_Norm'] == 'ANTICIPO-ISLR':
                veces = anticipos_por_mes.get((row['Anio_Cruce'], row['Mes_Cruce']), 1)
                if veces == 1: estatus_base += f" (Único pago en {row['Frecuencia']})"
            return estatus_base

    df_audit['Estatus_Fiscal'] = df_audit.apply(determinar_estatus, axis=1)

    df_audit['Días de Desviación'] = np.where(
        df_audit['Estatus_Fiscal'].str.contains('Exonerado'), 0,
        np.where(
            df_audit['_merge'] == 'left_only',
            (fecha_contingencia_dt - df_audit['fecha_limite']).dt.days,
            (df_audit['Fecha_Operacion_Real'] - df_audit['fecha_limite']).dt.days
        )
    )

    col_monto = [c for c in df_audit.columns if 'monto' in c.lower() or 'total' in c.lower()]
    if col_monto:
        df_audit['Monto(Bs.)'] = pd.to_numeric(df_audit[col_monto[0]], errors='coerce').fillna(0.0).round(2)
    else:
        df_audit['Monto(Bs.)'] = 0.0

    tasas_bcv = obtener_tasas_bcv()
    tasa_eur_data = tasas_bcv.get("EUR", {})
    tasa_calculo = round(float(tasa_eur_data["tasa"]), 2) if tasa_eur_data.get("tasa") else 0.0

    df_audit['Días_Mora_Real'] = np.where(df_audit['Días de Desviación'] > 0, df_audit['Días de Desviación'], 0)
    df_audit['Multa_Acumulada(Bs.)'] = (df_audit['Monto(Bs.)'] * 0.05 * df_audit['Días_Mora_Real']).round(2)
    df_audit['Multa_Art_103(EUR)'] = np.where(df_audit['Estatus_Fiscal'].str.contains('Pagado Tarde'), 100.0, 0.0)
    df_audit['Multa_Art_108(EUR)'] = np.where(
        df_audit['Estatus_Fiscal'].str.contains('Pagado Tarde') & (df_audit['Fase_Fiscal'].str.lower() == 'especial'),
        200.0, 0.0
    )
    df_audit['Multas_Fijas(EUR)'] = (df_audit['Multa_Art_103(EUR)'] + df_audit['Multa_Art_108(EUR)']).round(2)
    df_audit['Multas_Fijas(Bs.)'] = (df_audit['Multas_Fijas(EUR)'] * tasa_calculo).round(2)
    df_audit['Total_Deuda(Bs.)'] = np.where(
        df_audit['Estatus_Fiscal'].str.contains('NO PAGADO|Pendiente'),
        df_audit['Monto(Bs.)'] + df_audit['Multa_Acumulada(Bs.)'] + df_audit['Multas_Fijas(Bs.)'],
        df_audit['Multa_Acumulada(Bs.)'] + df_audit['Multas_Fijas(Bs.)']
    )
    df_audit['Total_Deuda(Bs.)'] = pd.to_numeric(df_audit['Total_Deuda(Bs.)'], errors='coerce').round(2)
    df_audit['Tasa_BCV_EUR'] = tasa_calculo
    df_audit['Deuda_Total(EUR)'] = (df_audit['Total_Deuda(Bs.)'] / tasa_calculo).round(2) if tasa_calculo > 0 else 0.0

    def display_impuesto(row):
        if row['Renta_Norm'] == 'ISLR-ANUAL': return 'ISLR (Declaración Anual)'
        if row['Renta_Norm'] == 'ISLR-RETENCION': return 'ISLR (Retención Mensual)'
        return row['Renta_Norm']

    df_audit['Impuesto'] = df_audit.apply(display_impuesto, axis=1)
    df_audit['Año'] = df_audit['Anio_Cruce'].astype(int)
    df_audit['Mes_Num_Final'] = df_audit['Mes_Cruce'].astype(int)

    df_audit['Concepto Contable'] = df_audit[nombre_concepto].fillna('— Sin Registro —') \
        if nombre_concepto in df_audit.columns else '— Sin Registro —'
    df_audit['Periodo Registrado en Excel'] = df_audit['Periodo_Original'].fillna('— No Pagado —') \
        if 'Periodo_Original' in df_audit.columns else '— No Pagado —'

    meses_nombres = {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
                     7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre',
                     12: 'Diciembre', 0: 'Desconocido', 99: 'Cierre Anual'}
    df_audit['Mes_Str'] = df_audit['Mes_Num_Final'].map(meses_nombres)
    df_audit['Periodo'] = df_audit['Mes_Str'] + ' ' + df_audit['Año'].astype(str)
    df_audit['Fecha Límite Oficial'] = df_audit['fecha_limite'].dt.strftime('%d/%m/%Y').fillna('Sin Fecha Oficial')
    df_audit['Fecha de Pago Real'] = df_audit['Fecha_Operacion_Real'].dt.strftime('%d/%m/%Y').fillna('— No Pagado —') \
        if 'Fecha_Operacion_Real' in df_audit.columns else '— No Pagado —'

    df_audit['Faltante'] = df_audit['Estatus_Fiscal'].str.contains('NO PAGADO|Pendiente')
    df_audit = df_audit.sort_values(by=['Faltante', 'Año', 'Mes_Num_Final', 'Impuesto', 'Frecuencia'])

    for col in ['Monto(Bs.)', 'Multa_Acumulada(Bs.)', 'Multa_Art_103(EUR)',
                'Multa_Art_108(EUR)', 'Multas_Fijas(EUR)', 'Multas_Fijas(Bs.)',
                'Total_Deuda(Bs.)', 'Tasa_BCV_EUR', 'Deuda_Total(EUR)']:
        if col in df_audit.columns:
            df_audit[col] = pd.to_numeric(df_audit[col], errors='coerce').round(2)

    columnas_finales = [
        'Fase_Fiscal', 'Periodo', 'Impuesto', 'Frecuencia', 'Concepto Contable',
        'Periodo Registrado en Excel', 'Fecha Límite Oficial', 'Fecha de Pago Real',
        'Monto(Bs.)', 'Estatus_Fiscal', 'Días de Desviación',
        'Multa_Acumulada(Bs.)', 'Multa_Art_103(EUR)', 'Multa_Art_108(EUR)',
        'Multas_Fijas(Bs.)', 'Total_Deuda(Bs.)', 'Tasa_BCV_EUR', 'Deuda_Total(EUR)'
    ]
    columnas_finales = [c for c in columnas_finales if c in df_audit.columns]
    df_reporte = df_audit[columnas_finales].rename(columns={'Días de Desviación': 'Días de Mora/Desviación'})

    info_expediente = {
        "digito": digito_usuario,
        "tipo": tipo_final_empresa,
        "rango_excel": periodo_str,
        "tipo_cierre": tipo_cierre_usuario,
        "duplicados_eliminados": duplicados_eliminados,
        "fecha_auditoria": fecha_contingencia_dt.strftime('%d/%m/%Y')
    }

    return df_reporte, df_cal_empresa, anios_presentes, info_expediente


# ============================================================
# EXPORTACIÓN EXCEL CON FORMATO #,##0.00
# ============================================================
def generar_excel_multitabla(df_detalles):
    COLS_NUM = ['Monto(Bs.)', 'Multa_Acumulada(Bs.)', 'Multa_Art_103(EUR)',
                'Multa_Art_108(EUR)', 'Multas_Fijas(Bs.)', 'Total_Deuda(Bs.)',
                'Tasa_BCV_EUR', 'Deuda_Total(EUR)']

    def aplicar_formato_excel(ws):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_cells = [cell.value for cell in ws[1]]
        header_fill = PatternFill("solid", fgColor="1D4ED8")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin = Side(style='thin', color="D1D5DB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        for col_name in COLS_NUM:
            if col_name in header_cells:
                col_idx = header_cells.index(col_name) + 1
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or '')))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)
        light_fill = PatternFill("solid", fgColor="F8FAFC")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                if row_idx % 2 == 0:
                    if not cell.fill or cell.fill.fill_type == 'none':
                        cell.fill = light_fill
                cell.border = border

    df_export = df_detalles.copy()
    for col in COLS_NUM:
        if col in df_export.columns:
            df_export[col] = pd.to_numeric(df_export[col], errors='coerce').round(2)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Auditoría Completa')
        aplicar_formato_excel(writer.sheets['Auditoría Completa'])

        for fase in df_export['Fase_Fiscal'].dropna().unique():
            sheet_name = f'Fase {fase}'
            df_export[df_export['Fase_Fiscal'] == fase].to_excel(writer, index=False, sheet_name=sheet_name)
            aplicar_formato_excel(writer.sheets[sheet_name])

        df_con_pago = df_export[
            df_export['Fecha de Pago Real'].notna() &
            (~df_export['Fecha de Pago Real'].astype(str).str.contains('No Pagado|—', na=False))
        ].copy()
        if not df_con_pago.empty:
            def diag_export(dias):
                if pd.isna(dias): return 'Sin dato'
                d = int(dias)
                if d > 30:   return f'Tardío grave (+{d} días)'
                elif d > 0:  return f'Tardío leve (+{d} días)'
                elif d == 0: return 'Exacto'
                else:        return f'Anticipado ({abs(d)} días antes)'

            cols_f = ['Fase_Fiscal', 'Periodo', 'Impuesto', 'Frecuencia',
                      'Fecha Límite Oficial', 'Fecha de Pago Real',
                      'Días de Mora/Desviación', 'Monto(Bs.)', 'Estatus_Fiscal']
            df_fechas_exp = df_con_pago[[c for c in cols_f if c in df_con_pago.columns]].copy()
            df_fechas_exp['Diagnóstico Puntualidad'] = df_fechas_exp['Días de Mora/Desviación'].apply(diag_export)
            df_fechas_exp.to_excel(writer, index=False, sheet_name='Auditoría de Fechas')
            aplicar_formato_excel(writer.sheets['Auditoría de Fechas'])

            fases_f = df_fechas_exp['Fase_Fiscal'].dropna().unique()
            if len(fases_f) > 1:
                for fase in fases_f:
                    sn = f'Fechas {fase}'
                    df_fechas_exp[df_fechas_exp['Fase_Fiscal'] == fase].to_excel(writer, index=False, sheet_name=sn)
                    aplicar_formato_excel(writer.sheets[sn])

    return output.getvalue()


# ============================================================
# INTERFAZ
# ============================================================
st.set_page_config(page_title="Auditoría de Contingencias Tributarias", page_icon="⚖️", layout="wide")

with st.sidebar:
    st.markdown("## ⚙️ Configuración")
    st.markdown("---")
    st.markdown("### 🏢 Contribuyente")
    digito_seleccionado = st.number_input("Dígito Final del RIF", min_value=0, max_value=9, value=0, step=1)
    tipo_cierre_seleccionado = st.selectbox("Tipo de Cierre ISLR", options=["Ordinario", "Irregular"], index=0)
    st.markdown("### 📅 Corte de Evaluación")
    fecha_calculo = st.date_input("Fecha de Cálculo de Mora", datetime.now().date())
    st.markdown("---")
    st.markdown(
        "<div style='font-size:0.7rem;color:#94a3b8;text-align:center;line-height:1.6;'>"
        "Auditoría Tributaria<br>COT Art. 103 &amp; 108</div>",
        unsafe_allow_html=True
    )

col_logo, col_titulo = st.columns([0.06, 0.94])
with col_logo:
    st.markdown("<div style='font-size:2.6rem;margin-top:0.25rem;'>⚖️</div>", unsafe_allow_html=True)
with col_titulo:
    st.markdown("# Auditoría de Contingencias Tributarias")
    st.markdown(
        "<p style='color:#64748b;font-size:0.88rem;margin-top:-0.5rem;'>"
        "Análisis de cumplimiento &nbsp;·&nbsp; Cálculo de mora &nbsp;·&nbsp; Multas COT Art. 103 y 108"
        "</p>", unsafe_allow_html=True
    )
st.markdown("---")

df_calendario = obtener_calendario_db()

# ── UPLOAD HERO ──────────────────────────────────────────────
archivo_subido = None

if True:
    st.markdown("""
    <div class="upload-hero">
        <div class="upload-hero-top">
            <div class="upload-hero-title">📂 Cargar Archivo de Operaciones</div>
            <p class="upload-hero-sub">Sube el libro de operaciones del contribuyente para iniciar el análisis fiscal</p>
        </div>
        <div class="upload-hero-body">
            <div class="upload-chips">
                <div class="upload-chip"><span>📄</span> Formato .xlsx</div>
                <div class="upload-chip"><span>📋</span> Encabezados en fila 4</div>
                <div class="upload-chip"><span>📅</span> Fechas DD/MM/YYYY</div>
                <div class="upload-chip"><span>🔒</span> Procesado localmente</div>
            </div>
    """, unsafe_allow_html=True)

    archivo_subido = st.file_uploader(
        "Arrastra tu archivo aquí o haz clic para seleccionarlo",
        type=["xlsx"],
        help="El archivo debe tener encabezados en la fila 4 (header=3)",
        label_visibility="visible"
    )

    st.markdown("</div></div>", unsafe_allow_html=True)

    if archivo_subido is None:
        st.markdown("""
        <div style="
            display: flex; gap: 1rem; flex-wrap: wrap;
            margin-top: 1.2rem; justify-content: center;
        ">
            <div style="
                background: #ffffff; border: 1px solid #e2e8f0; border-radius: 14px;
                padding: 1.2rem 1.6rem; flex: 1; min-width: 180px; max-width: 240px;
                box-shadow: 0 1px 4px rgba(0,0,0,0.05); text-align: center;
            ">
                <div style="font-size: 1.8rem; margin-bottom: 0.4rem;">🔍</div>
                <div style="font-size: 0.8rem; font-weight: 700; color: #1e293b; margin-bottom: 0.3rem;">Detección automática</div>
                <div style="font-size: 0.75rem; color: #64748b;">Identifica tipo de contribuyente y periodos fiscales</div>
            </div>
            <div style="
                background: #ffffff; border: 1px solid #e2e8f0; border-radius: 14px;
                padding: 1.2rem 1.6rem; flex: 1; min-width: 180px; max-width: 240px;
                box-shadow: 0 1px 4px rgba(0,0,0,0.05); text-align: center;
            ">
                <div style="font-size: 1.8rem; margin-bottom: 0.4rem;">⚖️</div>
                <div style="font-size: 0.8rem; font-weight: 700; color: #1e293b; margin-bottom: 0.3rem;">Cálculo de multas</div>
                <div style="font-size: 0.75rem; color: #64748b;">Art. 103 y 108 COT con tasa BCV en tiempo real</div>
            </div>
            <div style="
                background: #ffffff; border: 1px solid #e2e8f0; border-radius: 14px;
                padding: 1.2rem 1.6rem; flex: 1; min-width: 180px; max-width: 240px;
                box-shadow: 0 1px 4px rgba(0,0,0,0.05); text-align: center;
            ">
                <div style="font-size: 1.8rem; margin-bottom: 0.4rem;">📊</div>
                <div style="font-size: 0.8rem; font-weight: 700; color: #1e293b; margin-bottom: 0.3rem;">Reporte exportable</div>
                <div style="font-size: 0.75rem; color: #64748b;">Descarga el análisis completo en Excel formateado</div>
            </div>
        </div>
        """, unsafe_allow_html=True)


# ── COLUMNAS NUMÉRICAS CON FORMATO VISUAL ───────────────────
COLS_BS  = ['Monto(Bs.)', 'Multa_Acumulada(Bs.)', 'Multas_Fijas(Bs.)', 'Total_Deuda(Bs.)']
COLS_EUR = ['Multa_Art_103(EUR)', 'Multa_Art_108(EUR)', 'Multas_Fijas(EUR)', 'Deuda_Total(EUR)', 'Tasa_BCV_EUR']
COLS_INT = ['Días de Mora/Desviación']

def aplicar_formato_visual(df):
    fmt = {}
    for c in COLS_BS:
        if c in df.columns: fmt[c] = '{:,.2f}'
    for c in COLS_EUR:
        if c in df.columns: fmt[c] = '{:,.2f}'
    for c in COLS_INT:
        if c in df.columns: fmt[c] = '{:,.0f}'
    return df.style.format(fmt, na_rep='—')

def renderizar_tablero_fase(titulo_fase, emoji, df_reporte_fase, df_cal_fase, anios):
    st.markdown(f"### {emoji} {titulo_fase}")
    if df_reporte_fase.empty and df_cal_fase.empty:
        st.info("No hay obligaciones ni pagos en esta fase para el periodo evaluado.")
        return

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Omitidos / Vencidos", len(df_reporte_fase[df_reporte_fase['Estatus_Fiscal'].str.contains('NO PAGADO', na=False)]))
    c2.metric("Pagos Tardíos", len(df_reporte_fase[df_reporte_fase['Estatus_Fiscal'].str.contains('Pagado Tarde', na=False)]))
    c3.metric("A Tiempo / Exonerado", len(df_reporte_fase[df_reporte_fase['Estatus_Fiscal'].str.contains('Adelantado|Exacto|Exonerado', na=False)]))
    c4.metric("Pagos Extras", len(df_reporte_fase[df_reporte_fase['Estatus_Fiscal'].str.contains('Extra', na=False)]))

    st.markdown("##### ⏰ Pagos Realizados Fuera de Fecha")
    df_morosos = df_reporte_fase[df_reporte_fase['Estatus_Fiscal'].str.contains('Pagado Tarde', na=False)]
    if not df_morosos.empty:
        st.dataframe(df_morosos.groupby('Impuesto').size().reset_index(name='Total Atrasados'), hide_index=True, use_container_width=True)
    else:
        st.success("✅ No hay pagos realizados fuera de fecha en esta fase.")

    st.markdown(" ")
    impuestos = list(df_reporte_fase['Impuesto'].dropna().unique())
    pestanas = st.tabs(["📋 Vista Completa"] + [f"📂 {i}" for i in impuestos])

    def color_estatus(val):
        v = str(val)
        if 'NO PAGADO' in v: return 'background-color: #fef2f2; color: #991b1b; font-weight: 600'
        elif 'Tarde' in v: return 'background-color: #fffbeb; color: #92400e; font-weight: 600'
        elif any(x in v for x in ['Adelantado', 'Exacto', 'Exonerado']): return 'background-color: #f0fdf4; color: #14532d; font-weight: 600'
        elif 'Por Vencer' in v: return 'background-color: #eff6ff; color: #1e3a5f; font-weight: 600'
        return ''

    with pestanas[0]:
        st.dataframe(
            aplicar_formato_visual(df_reporte_fase).map(color_estatus, subset=['Estatus_Fiscal']),
            use_container_width=True, hide_index=True
        )
    for i, imp in enumerate(impuestos):
        with pestanas[i + 1]:
            df_imp = df_reporte_fase[df_reporte_fase['Impuesto'] == imp]
            st.dataframe(
                aplicar_formato_visual(df_imp).map(color_estatus, subset=['Estatus_Fiscal']),
                use_container_width=True, hide_index=True
            )

    st.markdown("#### 📅 Cumplimiento por Año")
    for anio in sorted(anios):
        df_cal_anio = df_cal_fase[df_cal_fase['Anio_Cruce'] == anio]
        df_rep_anio = df_reporte_fase[df_reporte_fase['Periodo'].str.endswith(str(anio), na=False)]
        esperados = len(df_cal_anio)
        if esperados > 0:
            omitidos = df_rep_anio[df_rep_anio['Estatus_Fiscal'].str.contains('NO PAGADO|Pendiente', na=False)]
            n_omitidos = len(omitidos)
            icono = "🟢" if n_omitidos == 0 else "🔴"
            with st.expander(f"{icono} {anio}  ·  Exigidas: {esperados}  ·  Cumplidas: {esperados - n_omitidos}  ·  Pendientes: {n_omitidos}"):
                if n_omitidos > 0:
                    st.dataframe(
                        omitidos.groupby(['Impuesto', 'Frecuencia']).size().reset_index(name='Declaraciones Pendientes'),
                        hide_index=True, use_container_width=True
                    )
                else:
                    st.success(f"✅ Todas las {esperados} obligaciones cumplidas al corte.")


if archivo_subido is not None and not df_calendario.empty:
    with st.spinner("⚙️ Procesando auditoría tributaria..."):
        df_empresa = pd.read_excel(archivo_subido, header=3)
        df_reporte, df_cal_esperado, anios, info = procesar_auditoria_completa(
            df_empresa, df_calendario, digito_seleccionado, fecha_calculo, tipo_cierre_seleccionado
        )

    if info['duplicados_eliminados'] > 0:
        st.warning(f"🧹 **Limpieza automática:** {info['duplicados_eliminados']} registros duplicados eliminados.")

    st.markdown("## 📋 Expediente de Auditoría")
    st.markdown(f"""
    <div class="expediente-header">
        <div><div class="exp-label">📅 Corte</div><div class="exp-value">{info['fecha_auditoria']}</div></div>
        <div><div class="exp-label">📊 Periodo Excel</div><div class="exp-value">{info['rango_excel']}</div></div>
        <div><div class="exp-label">🏢 Dígito RIF</div><div class="exp-value">{info['digito']}</div></div>
        <div><div class="exp-label">📝 Estatus Fiscal</div><div class="exp-value">{info['tipo']}</div></div>
        <div><div class="exp-label">⚙️ Tipo Cierre</div><div class="exp-value">{info['tipo_cierre']}</div></div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("---")
    fases_presentes = set(df_reporte['Fase_Fiscal'].dropna().unique())

    if 'Ordinario' in fases_presentes and 'Especial' in fases_presentes:
        st.warning("⚠️ **Transición detectada:** El contribuyente migró de Ordinario a Especial.")
        tab_ord, tab_esp = st.tabs(["📘 Fase Ordinaria", "📙 Fase Especial"])
        with tab_ord:
            renderizar_tablero_fase("Fase Ordinaria", "📘",
                df_reporte[df_reporte['Fase_Fiscal'] == 'Ordinario'],
                df_cal_esperado[df_cal_esperado['Tipo_Contribuyente_Esperado'] == 'ordinario'], anios)
        with tab_esp:
            renderizar_tablero_fase("Fase Especial", "📙",
                df_reporte[df_reporte['Fase_Fiscal'] == 'Especial'],
                df_cal_esperado[df_cal_esperado['Tipo_Contribuyente_Esperado'] == 'especial'], anios)
    else:
        fase_unica = list(fases_presentes)[0] if fases_presentes else "Desconocida"
        renderizar_tablero_fase(f"Contribuyente {fase_unica}", "📊", df_reporte, df_cal_esperado, anios)

    st.markdown("---")

    # ── AUDITORÍA DE FECHAS ──────────────────────────────────
    st.markdown("### 📆 Auditoría de Fechas de Pago")
    st.markdown(
        "<p style='color:#64748b;font-size:0.88rem;margin-top:-0.5rem;margin-bottom:1rem;'>"
        "Comparativo entre la fecha en que el contribuyente realizó cada pago "
        "y la fecha límite oficial establecida. Solo se muestran registros con fecha de pago registrada."
        "</p>", unsafe_allow_html=True
    )

    df_con_pago = df_reporte[
        df_reporte['Fecha de Pago Real'].notna() &
        (~df_reporte['Fecha de Pago Real'].astype(str).str.contains('No Pagado|—', na=False))
    ].copy()

    if not df_con_pago.empty:
        df_fechas = df_con_pago[[
            'Fase_Fiscal', 'Periodo', 'Impuesto', 'Frecuencia',
            'Fecha Límite Oficial', 'Fecha de Pago Real',
            'Días de Mora/Desviación', 'Estatus_Fiscal', 'Monto(Bs.)'
        ]].copy()

        df_fechas['_limite_dt'] = pd.to_datetime(df_fechas['Fecha Límite Oficial'], format='%d/%m/%Y', errors='coerce')
        df_fechas['_pago_dt']   = pd.to_datetime(df_fechas['Fecha de Pago Real'],  format='%d/%m/%Y', errors='coerce')

        def diagnostico_fecha(row):
            dias = row['Días de Mora/Desviación']
            if pd.isna(dias): return '⚪ Sin dato'
            dias = int(dias)
            if dias > 30:  return f'🔴 Tardío grave (+{dias} días)'
            elif dias > 0: return f'🟠 Tardío leve (+{dias} días)'
            elif dias == 0: return '🔵 Exacto (mismo día)'
            else:           return f'🟢 Anticipado ({abs(dias)} días antes)'

        df_fechas['Diagnóstico'] = df_fechas.apply(diagnostico_fecha, axis=1)
        df_fechas['_anio'] = df_fechas['_limite_dt'].dt.year.fillna(0).astype(int)

        fases_audit = sorted(df_fechas['Fase_Fiscal'].dropna().unique())
        tiene_transicion = len(fases_audit) > 1

        def render_tabla_fechas(df_f, titulo):
            st.markdown(f"##### {titulo}")
            total   = len(df_f)
            tarde   = len(df_f[df_f['Días de Mora/Desviación'] > 0])
            exacto  = len(df_f[df_f['Días de Mora/Desviación'] == 0])
            adelant = len(df_f[df_f['Días de Mora/Desviación'] < 0])
            pct_ok  = round((exacto + adelant) / total * 100, 1) if total > 0 else 0

            m1, m2, m3, m4, m5 = st.columns(5)
            m1.metric("Total Pagos", total)
            m2.metric("🔴 Tardíos", tarde)
            m3.metric("🔵 Exactos", exacto)
            m4.metric("🟢 Anticipados", adelant)
            m5.metric("✅ Cumplimiento", f"{pct_ok}%")

            cols_mostrar = ['Periodo', 'Impuesto', 'Frecuencia',
                            'Fecha Límite Oficial', 'Fecha de Pago Real',
                            'Días de Mora/Desviación', 'Diagnóstico', 'Monto(Bs.)']
            df_vista = df_f[cols_mostrar].copy()

            def color_diag(val):
                v = str(val)
                if 'grave' in v:        return 'background-color:#fef2f2;color:#991b1b;font-weight:600'
                elif 'leve' in v:       return 'background-color:#fffbeb;color:#92400e;font-weight:600'
                elif 'Exacto' in v:     return 'background-color:#eff6ff;color:#1e3a5f;font-weight:600'
                elif 'Anticipado' in v: return 'background-color:#f0fdf4;color:#14532d;font-weight:600'
                return ''

            fmt_fechas = {'Días de Mora/Desviación': '{:,.0f}', 'Monto(Bs.)': '{:,.2f}'}
            st.dataframe(
                df_vista.style.format(fmt_fechas, na_rep='—').map(color_diag, subset=['Diagnóstico']),
                use_container_width=True, hide_index=True
            )

            st.markdown("###### 📅 Detalle por Año")
            for anio in sorted(df_f['_anio'].unique()):
                if anio == 0: continue
                df_a = df_f[df_f['_anio'] == anio]
                n_tarde = len(df_a[df_a['Días de Mora/Desviación'] > 0])
                icono = "🟢" if n_tarde == 0 else "🔴"
                with st.expander(f"{icono} {anio}  ·  {len(df_a)} pagos  ·  {n_tarde} tardíos"):
                    df_a_vista = df_a[cols_mostrar].copy()
                    st.dataframe(
                        df_a_vista.style.format(fmt_fechas, na_rep='—').map(color_diag, subset=['Diagnóstico']),
                        use_container_width=True, hide_index=True
                    )

        if tiene_transicion:
            tab_ord_f, tab_esp_f, tab_total_f = st.tabs([
                "📘 Ordinario — Fechas", "📙 Especial — Fechas", "🔢 Conjunto Total"
            ])
            with tab_ord_f:
                render_tabla_fechas(df_fechas[df_fechas['Fase_Fiscal'] == 'Ordinario'], "Fase Ordinaria — Comparativo de Fechas")
            with tab_esp_f:
                render_tabla_fechas(df_fechas[df_fechas['Fase_Fiscal'] == 'Especial'], "Fase Especial — Comparativo de Fechas")
            with tab_total_f:
                render_tabla_fechas(df_fechas, "Conjunto Total (Ordinario + Especial)")
        else:
            render_tabla_fechas(df_fechas, "Comparativo de Fechas de Pago")
    else:
        st.info("No se encontraron registros con fecha de pago registrada para construir el comparativo.")

    st.markdown("---")
    st.markdown("### 📊 Resumen Global de Contingencia")

    tasas_bcv_display = obtener_tasas_bcv()
    eur_d = tasas_bcv_display.get("EUR", {})
    usd_d = tasas_bcv_display.get("USD", {})
    tasa_oficial = eur_d.get("tasa")
    tasa_usd_oficial = usd_d.get("tasa")

    bcv_cols = st.columns(2)
    for col_ui, moneda_key, moneda_data, icono_vivo in [
        (bcv_cols[0], "EUR", eur_d, "💶"),
        (bcv_cols[1], "USD", usd_d, "💵"),
    ]:
        t = moneda_data.get("tasa")
        f = moneda_data.get("fecha", "—")
        cache = moneda_data.get("es_cache", False) or (f and "caché BD" in f)
        if t:
            badge_class = "bcv-badge-cache" if cache else "bcv-badge"
            icono = "🗄️" if cache else icono_vivo
            col_ui.markdown(
                f'<div class="{badge_class}">{icono} Tasa BCV {moneda_key}: '
                f'<strong>Bs.&nbsp;{t:,.2f}</strong>&nbsp;&nbsp;·&nbsp;&nbsp;{f}</div>',
                unsafe_allow_html=True
            )

    ck1, ck2, ck3 = st.columns(3)
    ck1.metric("Omitidos / Vencidos ❌", len(df_reporte[df_reporte['Estatus_Fiscal'].str.contains('NO PAGADO', na=False)]))
    ck2.metric("Pagos Tardíos 🔴", len(df_reporte[df_reporte['Estatus_Fiscal'].str.contains('Pagado Tarde', na=False)]))
    ck3.metric("Pagos Extras ⚠️", len(df_reporte[df_reporte['Estatus_Fiscal'].str.contains('Extra', na=False)]))

    st.markdown("#### 💰 Impacto Financiero Estimado (COT)")
    tasa_uso = tasa_oficial if tasa_oficial else 1.0

    def construir_desglose(df_fuente, tasa):
        monto_p = df_fuente[df_fuente['Estatus_Fiscal'].str.contains('NO PAGADO|Pendiente')]['Monto(Bs.)'].sum()
        mora    = df_fuente['Multa_Acumulada(Bs.)'].sum()
        a103    = df_fuente['Multa_Art_103(EUR)'].sum()
        a108    = df_fuente['Multa_Art_108(EUR)'].sum()
        tot_bs  = df_fuente['Total_Deuda(Bs.)'].sum()
        tot_eur = pd.to_numeric(df_fuente['Deuda_Total(EUR)'], errors='coerce').sum()
        return pd.DataFrame({
            "Concepto": [
                "🏛️ Deuda Principal (Impuestos Omitidos)",
                "📈 Multas por Mora (5% diario s/monto)",
                "📄 COT Art. 103 — 100 EUR por pago tardío",
                "🏢 COT Art. 108 — 200 EUR adicional (Especiales)",
                "🚨 TOTAL RIESGO ESTIMADO"
            ],
            "Monto (Bs.)": [
                f"Bs. {monto_p:,.2f}", f"Bs. {mora:,.2f}",
                f"Bs. {a103 * tasa:,.2f}", f"Bs. {a108 * tasa:,.2f}",
                f"Bs. {tot_bs:,.2f}"
            ],
            "Equivalente (EUR)": [
                f"€ {monto_p / tasa:,.2f}" if tasa_oficial else "—",
                f"€ {mora / tasa:,.2f}"    if tasa_oficial else "—",
                f"€ {a103:,.2f}", f"€ {a108:,.2f}", f"€ {tot_eur:,.2f}"
            ]
        })

    if 'Ordinario' in fases_presentes and 'Especial' in fases_presentes:
        tab_res_ord, tab_res_esp, tab_res_tot = st.tabs(["📘 Ordinario", "📙 Especial", "🔢 Total Conjunto"])
        with tab_res_ord:
            st.dataframe(construir_desglose(df_reporte[df_reporte['Fase_Fiscal'] == 'Ordinario'], tasa_uso), hide_index=True, use_container_width=True)
        with tab_res_esp:
            st.dataframe(construir_desglose(df_reporte[df_reporte['Fase_Fiscal'] == 'Especial'], tasa_uso), hide_index=True, use_container_width=True)
        with tab_res_tot:
            st.dataframe(construir_desglose(df_reporte, tasa_uso), hide_index=True, use_container_width=True)
    else:
        st.dataframe(construir_desglose(df_reporte, tasa_uso), hide_index=True, use_container_width=True)

    st.markdown("---")
    excel_data = generar_excel_multitabla(df_reporte)
    st.download_button(
        label="📥 Descargar Reporte Completo (.xlsx)", data=excel_data,
        file_name=f"Auditoria_Tributaria_{info['digito']}_{info['fecha_auditoria'].replace('/', '-')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="btn_descargar_auditoria"
    )

elif df_calendario.empty:
    st.error("⚠️ No se pudo cargar el calendario tributario desde la base de datos.")